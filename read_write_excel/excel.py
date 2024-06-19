import logging

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

import config
from model.models import Tweet, Reply

path = config.excel_file_path
replies_path = config.replies_excel_file_path
tweets_path = config.tweets_excel_file_path

username_column = config.excel_username_column


def get_usernames_from_excel():
    df = pd.read_excel(path)

    # Check if the usernames are with the @ at the beginning
    aux = []
    usernames = df[username_column]

    for username in usernames:
        if username[0] == "@":
            aux.append(username[1:])
        else:
            aux.append(username)

    return aux


def write_followers_to_excel(users_info: [{}], first_position: int):
    df = pd.read_excel(path)

    i = first_position
    for user in users_info:
        df["Followers"][i] = user["followers_count"]
        df["Tweet count"][i] = user["tweet_count"]
        i += 1

    df.to_excel(path)


def save_tweets_to_excel_by_date(tweets: [Tweet]):
    id_list = []
    text_list = []

    for tweet in tweets:
        id_list.append(tweet.tweet_id)
        text_list.append(tweet.text)

    # TODO: fix some issues with categories and labels in columns

    df = pd.DataFrame({"Tweet id": id_list, "Tweet": text_list})

    df.to_excel(tweets_path)

    # Test of creating dropdown list
    wb = load_workbook(filename=tweets_path)
    ws = wb.active

    dv = DataValidation(type="list", formula1='"Prueba 1, Prueba 2"', allow_blank=False)

    dv.error = 'Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    aux = "D2:D" + str(len(tweets))

    dv.showInputMessage = True
    dv.showErrorMessage = True

    dv.add(aux)

    ws.add_data_validation(dv)

    wb.save(tweets_path)

    logging.info(f"Tweets saved to {tweets_path}")


def save_tweets_to_excel(tweets: [Tweet]):
    id_list = []
    text_list = []

    for tweet in tweets:
        id_list.append(tweet.tweet_id)
        text_list.append(tweet.text)

    # TODO: fix some issues with categories and labels in columns

    df = pd.DataFrame({"Tweet id": id_list, "Tweet": text_list})

    df.to_excel(tweets_path)

    # Test of creating dropdown list
    wb = load_workbook(filename=tweets_path)
    ws = wb.active

    ws.cell(row=1, column=4).value = "Theme"

    dv = DataValidation(type="list",
                        formula1='"Immigration, Gender, Ukraine, Israel/Palestina, EU Elections, Sports, National Politics, Climate Change, Economy and Finance, Healthcare, International Politics, Culture, Science and Technology"',
                        allow_blank=False)

    dv.error = 'Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    aux = "D2:D" + str(len(tweets) + 1)

    dv.showInputMessage = True
    dv.showErrorMessage = True

    dv.add(aux)

    ws.add_data_validation(dv)

    wb.save(tweets_path)

    logging.info(f"Tweets saved to {tweets_path}")


def save_replies_to_excel(replies: [Reply]):
    id_list = []
    text_list = []
    conversation_list = []

    for tweet in replies:
        id_list.append(tweet.tweet_id)
        text_list.append(tweet.text)
        conversation_list.append(tweet.conversation_id)

    # TODO: fix some issues with categories and labels in columns

    df = pd.DataFrame({"Tweet id": id_list, "Conversation id": conversation_list, "Tweet": text_list})

    df.to_excel(replies_path)

    # Test of creating dropdown list
    wb = load_workbook(filename=replies_path)
    ws = wb.active

    ws.cell(row=1, column=5).value = "Contains hate"
    ws.cell(row=1, column=6).value = "Hate type"

    dv = DataValidation(type="list", formula1='"0, 1"', allow_blank=False)

    dv.error = 'Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    aux = "E2:E" + str(len(replies) + 1)

    dv.showInputMessage = True
    dv.showErrorMessage = True

    dv.add(aux)

    ws.add_data_validation(dv)

    dv1 = DataValidation(type="list", formula1='"0, 1, 2, 3"', allow_blank=False)

    dv1.error = 'Your entry is not in the list'
    dv1.errorTitle = 'Invalid Entry'

    dv1.prompt = 'Please select from the list'
    dv1.promptTitle = 'List Selection'

    aux = "F2:F" + str(len(replies) + 1)

    dv1.showInputMessage = True
    dv1.showErrorMessage = True

    dv1.add(aux)

    ws.add_data_validation(dv1)

    wb.save(replies_path)

    logging.info(f"Tweets saved to {replies_path}")
